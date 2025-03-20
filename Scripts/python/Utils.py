import openpyxl


class Utils:
    def write_list_to_file(self, data_list: list, file_path: str):
        with open(file_path, "w", encoding="UTF-8")  as file_list:
            for line in data_list:
                file_list.write(str(line) + "\n")
            print("File created: " + file_path)


    def write_content_to_file(self, file_path: str, content: str):
        with open(file_path, 'w', encoding="UTF-8") as file_content:
            file_content.write(content)


    def get_data_from_xlsx_file(self, filename: str, column_list :list):
        """Return data in columns from column_list throught the yield.

        Args:
            filename (str): filename (Exapmle: imgay.xlsx)
            column_list (list): list of colums (Example: [1, 2, 4, 6])

        yeild: 
            result (list): rows from each column in column_list
        """
        workbook = openpyxl.load_workbook(filename=filename)
        sheet = workbook.active

        for column in column_list:
            result = []
            for row in range(1, sheet.max_row):
                result.append(sheet.cell(row, column).value)
            yield result