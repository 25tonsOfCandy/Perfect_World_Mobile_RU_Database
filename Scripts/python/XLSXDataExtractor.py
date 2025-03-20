import openpyxl


class XLSXDataExtractor:
    def extract_data(self, filename: str, column_list :list):
        """
        yeild: 
            result (list): rows from each column in column_list
        """
        workbook = openpyxl.load_workbook(filename=filename)
        sheet = workbook.active

        for column in column_list:
            result = []
            for row in range(1, sheet.max_row):
                result.append(sheet.cell(row, column).value)
            yield result