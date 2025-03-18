"""
"""
from CardGenerator import CardGenerator
import openpyxl
import os
# PRESETS
# BESTIARY - for generating bestiary cards
# BESTIARY_SETS - for generating bestiary sets pages
# GLYPHS - for generating glyph cards
# BESTIARY_PARTS - for generating bestiary parts
# NULL ("") - do nothing
PRESET = "BESTIARY_PARTS"
DIRECTORIES = ['bestiary', 'bestiary_sets', 'glyphs'] # Important directories for checking


# UTILS
def get_data_from_xlsx_file(filename: str, column_list :list):
    """Return data in columns from column_list throught the yield.

    Args:
        filename (str): filename (Exapmle: imgay.xlsx)
        column_list (list): list of colums (Example: [1, 2, 4, 6])

    yeild: 
        result (list): rows from each column in column_list
    """
    workbook = openpyxl.load_workbook(filename=filename)
    sheet = workbook.active

    for column in column_list:
        result = []
        for row in range(1, sheet.max_row):
            result.append(sheet.cell(row, column).value)
        yield result


def write_list_to_file(data_list: list, path_to_file_with_filename: str):
    with open(path_to_file_with_filename, "w", encoding="UTF-8")  as file:
        for line in data_list:
            file.write(str(line) + "\n")
        print("File created: " + path_to_file_with_filename)


def check_exists_all_important_directories():
    for directory in DIRECTORIES:
        if not os.path.exists(directory):
            os.makedirs(directory)


def main():
    check_exists_all_important_directories()
    generator = CardGenerator()
    if PRESET == "BESTIARY":
        raw_data = get_data_from_xlsx_file(filename = 'Bestiary_Skills_Source_RUCHNames.xlsx', column_list = [3, 4, 5])
        names_list = next(raw_data)
        description_list = next(raw_data)
        source_list = next(raw_data)

        generator.generate_bestiary_cards(names_list, source_list, description_list)
        return 0


    elif PRESET == "BESTIARY_SETS":
        raw_data = get_data_from_xlsx_file(filename = 'BestiarySetBonuses.xlsx', column_list = [1, 2, 3])
        sets_list = next(raw_data)
        sets_bonus_list = next(raw_data)
        sets_bonus_value_list = next(raw_data)

        generator.generate_bestiary_sets(sets_list, sets_bonus_list, sets_bonus_value_list)
        return 0


    elif PRESET == "GLYPHS":
        raw_data = get_data_from_xlsx_file(filename = 'Glyphs.xlsx', column_list = [2, 3, 4, 5])
        glyph_names_list = next(raw_data)
        glyph_description_list = next(raw_data)
        glyph_stats_list = next(raw_data)
        glyph_stat_values_list = next(raw_data)

        generator.generate_glyph_cards(glyph_names_list, glyph_description_list, glyph_stats_list, glyph_stat_values_list)
        return 0

    elif PRESET == "BESTIARY_PARTS":
        raw_data = get_data_from_xlsx_file(filename='BestiaryParts.xlsx', column_list=[1, 2])
        bestiary_part_names_list = next(raw_data)
        bestiary_part_descriptions_list = next(raw_data)
        generator.generate_bestiary_part_cards(bestiary_part_names_list, bestiary_part_descriptions_list)
        return 0

    if PRESET == "":
        return 1
    else:
        return 2


if __name__ == '__main__':
    result = main()
    if result == 1:
        print('Preset is empty')

    elif result == 2:
        print('Wrong preset')

    else: print("Complete!")
    
